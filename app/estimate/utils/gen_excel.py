from django.conf import settings
import os
import datetime
import json
from pytz import timezone
from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font


# load workbook worksheet
def load_worksheet():
    template_path = os.path.join(settings.STATIC_ROOT, 'templates')
    template_name = "template_estimate.xlsx"
    wb = load_workbook(f'{template_path}/{template_name}')
    ws = wb['견적서']
    return wb, ws


# excel data edit
def create_report(claimant, estimate_doc, estimate_num):
    returncode = 0
    msg = ""

    # data init
    now_time = datetime.datetime.now(timezone('Asia/Seoul')).strftime('%Y-%m-%d')
    project_name = estimate_doc["project_name"]
    project_scope = json.loads(estimate_doc["project_scope"])
    mm_result = estimate_doc["mm_result"]
    company_name = claimant["company_name"]
    email = claimant["email"].replace("@", "-")

    # is file exists
    doc_path = os.path.join(settings.MEDIA_ROOT, 'result_estimate_doc')
    # 디렉터리 생성: 이미 있으면 무시
    os.makedirs(doc_path, exist_ok=True)

    report_name = f"{company_name}_{email}_{now_time}"
    report_path = f'{doc_path}/{report_name}.xlsx'
    if os.path.exists(report_path):
        return returncode, report_name, report_path

    # load workbook worksheet
    wb, ws = load_worksheet()

    # insert doc data
    ws["A6"].value = f"{company_name} 귀중"
    ws["A6"].font = Font(size=14)

    # 견적번호: A23000부터 순차 증가
    ws["B9"].value = estimate_num
    # 견적일자
    ws["B10"].value = now_time

    ws["A15"].value = f"1. 사업명: {project_name}"
    ws["A16"].value = f"""2. 업무범위
    가. 점검대상: {", ".join(project_scope)}
    나. 개인정보 기술적 보호조치 점검
    다. 취약점 조치 확인 이행점검\
    """
    ws["G20"].value = mm_result

    # file save
    wb.save(report_path)
    return returncode, report_name, report_path
